import os
import re
from datetime import datetime
import csv

from flask import Flask, flash, redirect, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = 'Qd5vX276MsEtH7EhiMAFyqGAk9QV2tC7'

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def index():
    return render_template('index.html')

@app.errorhandler(500)
def server_error(error):
    # render the error template with the traceback
    return render_template('error.html', error=error), 500



@app.route('/process', methods=['GET','POST'])
def process():
    
    if 'file' not in request.files:
        print("NO FILES")
    # Get the uploaded files
    file1 = request.files['file1']
    file2 = request.files['file2']
    
    if not allowed_file(file1.filename) or not allowed_file(file2.filename):
        flash('Please make sure the attached files are correct.', 'error')
        return redirect("/")
    
    # Save the files to disk
    currentTime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    file1.save(os.path.join("uploads/missingEVV", currentTime + "_" + "missingEVV.xlsx"))
    file2.save(os.path.join("uploads/caregivers", currentTime + "_" + "caregiver_list.xlsx"))


    # Define the regex pattern for the time column
    date_pattern = re.compile(r'^(0?[1-9]|1[012])/(0?[1-9]|[12][0-9]|3[01])/\d{4}$')

    # Define the regex pattern for the string column
    aidecode_pattern = re.compile(r'^(GTH|128)-\d{4}$')
    aidecode_four_digit_pattern = re.compile(r'^\d{4}$')
    
    #Define the regex pattern for xxx-xxx-xxxx phone number format
    phone_number_pattern = re.compile(r'^\d{3}-\d{3}-\d{4}$')

    # Define the number of rows to search
    num_rows = 10

    # Initialize variables to store the column letters
    missingEVV_date_column = None
    missingEVV_aidecode_column = None
    missingEVV_worksheet_index = None

    # Load the Excel files and process the data (replace with your own processing code)
    wb_missing = load_workbook(os.path.join("uploads/missingEVV", currentTime + "_" + "missingEVV.xlsx"))
    
    # Loop through each worksheet in the workbook
    for index, worksheet in enumerate(wb_missing):
        # Loop through each column in the worksheet
        for column in worksheet.columns:
            # Loop through each cell in the first 'num_rows' rows of the column
            for cell in column[:num_rows]:
                # Check if the cell value matches the date regex pattern
                if cell.value and date_pattern.match(str(cell.value)):
                    # Store the column letter in the 'date_column' variable
                    missingEVV_date_column = cell.column_letter
                # Check if the cell value matches the string regex pattern
                elif cell.value and (aidecode_pattern.match(str(cell.value)) or aidecode_four_digit_pattern.match(str(cell.value))):
                    # Store the column letter in the 'missingEVV_aidecode_column' variable
                    missingEVV_aidecode_column = cell.column_letter
            # If both column letters have been found, exit the loops
            if missingEVV_date_column and missingEVV_aidecode_column:
                break
        # If both column letters have been found, exit the loops
        if missingEVV_date_column and missingEVV_aidecode_column:
            missingEVV_worksheet_index = index
            break

    # Print the results
    # if missingEVV_date_column and missingEVV_aidecode_column:
    #     print(f"Date column found in column {missingEVV_date_column}")
    #     print(f"String column found in column {missingEVV_aidecode_column}")
    #     print(f"Worksheet index is {missingEVV_worksheet_index}")
    # else:
    #     print("Could not find both date and string columns")

    ws_missing = wb_missing.worksheets[missingEVV_worksheet_index]
    

    # Initialize variables to store the column letters
    caregiver_aidecode_column = None
    caregiver_phone_number_column = None
    caregiver_worksheet_index = None
    
    wb_caregiver = load_workbook(os.path.join("uploads/caregivers", currentTime + "_" + "caregiver_list.xlsx"))
    
    # Loop through each worksheet in the workbook
    for index, worksheet in enumerate(wb_caregiver):
        # Loop through each column in the worksheet
        for column in worksheet.columns:
            # Loop through each cell in the first 'num_rows' rows of the column
            for cell in column[:num_rows]:
                # Check if the cell value matches the date regex pattern
                if cell.value and phone_number_pattern.match(str(cell.value)):
                    # Store the column letter in the 'caregiver_aidecode_column' variable
                    caregiver_phone_number_column = cell.column_letter
                # Check if the cell value matches the string regex pattern
                elif cell.value and (aidecode_pattern.match(str(cell.value)) or aidecode_four_digit_pattern.match(str(cell.value))):
                    # Store the column letter in the 'missingEVV_aidecode_column' variable
                    caregiver_aidecode_column = cell.column_letter
            # If both column letters have been found, exit the loops
            if caregiver_phone_number_column and caregiver_aidecode_column:
                break
        # If both column letters have been found, exit the loops
        if caregiver_phone_number_column and caregiver_aidecode_column:
            caregiver_worksheet_index = index
            break
        
        
    # Print the results
    # if missingEVV_date_column and missingEVV_aidecode_column:
    #     print(f"Phone number column found in column {caregiver_phone_number_column}")
    #     print(f"Aidecode column found in column {caregiver_aidecode_column}")
    #     print(f"Caregiver Worksheet index is {caregiver_worksheet_index}")
    # else:
    #     print("Could not find both date and string columns")
    
    ws_caregiver = wb_caregiver.worksheets[caregiver_worksheet_index]
    
##!@CONVERT LETTER TO INDEX FOR COLUMNS!!!

    
    new_wb = Workbook()
    new_ws = new_wb.active
    
    new_ws['A1'] = 'Mobile Number'
    new_ws['B1'] = 'Date'
    new_ws['C1'] = 'Aide'
# Create a dictionary to store the dates for each person
    dates_dict = {}

    # Loop through each row of the sheet
    for row in ws_missing.iter_rows(min_row=2, values_only=True):
        aide_code = str(row[column_index_from_string(missingEVV_aidecode_column)-1])[-4:] 
        date = row[column_index_from_string(missingEVV_date_column)-1] 
        # Add the date to the dictionary for this person
        if aide_code in dates_dict:
            if date not in dates_dict[aide_code]:
                dates_dict[aide_code].append(date)  
        else:
            dates_dict[aide_code] = [date]

    # Loop through the dictionary and update the sheet
    row_num = 2
    for aide_code, dates in dates_dict.items():
        # Write the name to column A
        new_ws.cell(row=row_num, column=2, value=", ".join(sorted(dates)))   
        new_ws.cell(row=row_num, column=3, value=aide_code)
        row_num += 1
        
        
        
    number_dict = {}
    
    for row in ws_caregiver.iter_rows(min_row=2, values_only=True):
        aide_code = str(row[column_index_from_string(caregiver_aidecode_column)-1])[-4:] #column B
        if row[1] != "":
            phone_number = "+1" + str(row[column_index_from_string(caregiver_phone_number_column)-1]).replace("-", "") # column C
            number_dict[aide_code] = [phone_number]
       
    
    for col in new_ws.iter_cols(min_col=3, max_col=3):
        for cell in col:
            if cell.value in number_dict:
             new_ws.cell(row=cell.row, column=1, value=number_dict[cell.value][0]) 
    

    # Save the result workbook to disk
    output_filename = currentTime + "_output.xlsx"
    new_wb.save(os.path.join("uploads/output", output_filename))
    
    with open('data.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        for row in new_ws.iter_rows():
            writer.writerow([cell.value for cell in row])

    file.close()
    
    response =  send_file("data.csv", attachment_filename='output.csv', as_attachment=True)
    
    os.remove("data.csv")
    
    return response

if __name__ == '__main__':
    app.run(debug=False, port=80, host="0.0.0.0")
